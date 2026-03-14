"""
Ejercicio 2 - Portafolios
Prueba Técnica Analista Jr de Inversiones - Insights WM
=======================================================

2.1 Fórmulas en términos de P y R:
    - Retorno esperado del portafolio:   E[Rp] = P · mean(R)     (producto punto)
    - Volatilidad esperada del portafolio: σp = sqrt(P · Cov(R) · Pᵀ)

2.2 Cómputo numérico para P1 y P2
"""

import numpy as np
import openpyxl
import time

# ──────────────────────────────────────────────
# 1. Carga de datos
# ──────────────────────────────────────────────
print("Cargando datos del archivo prueba.xlsx ...")
t0 = time.time()

wb = openpyxl.load_workbook(
    "/Users/santiagofajardoavendano/Downloads/prueba.xlsx",
    read_only=True,
    data_only=True
)

# ── Pesos P1 y P2 ──
def load_weights(wb, sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    return np.array(rows[1], dtype=float)

P1 = load_weights(wb, "P1")
P2 = load_weights(wb, "P2")

# ── Matriz de retornos R [30000 x 55] ──
ws_r = wb["Matriz de Simulacion"]
data = []
for i, row in enumerate(ws_r.iter_rows(min_row=2, values_only=True)):  # skip header
    data.append(row)

R = np.array(data, dtype=float)   # shape (30000, 55)

print(f"  Datos cargados en {time.time()-t0:.1f}s")
print(f"  Dimensiones R : {R.shape}")
print(f"  Dimensiones P1: {P1.shape}  |  suma = {P1.sum():.4f}")
print(f"  Dimensiones P2: {P2.shape}  |  suma = {P2.sum():.4f}")

# ──────────────────────────────────────────────
# 2.1  FÓRMULAS EN TÉRMINOS DE P y R
# ──────────────────────────────────────────────
print("\n" + "="*60)
print("2.1  FÓRMULAS (en términos de P y R)")
print("="*60)
print("""
Sea:
  R  : matriz de retornos simulados de dimensión [N x M]
       donde N = 30,000 escenarios y M = 55 activos.
  P  : vector de pesos del portafolio de dimensión [1 x M]
       (P1 o P2), donde sum(P) = 1.

── Retorno esperado del portafolio ──────────────────────────
  μ_activos = mean(R, axis=0)          → vector [M x 1]
                                         (retorno medio de cada activo)

  E[Rp] = P · μ_activos               → escalar
          = Σ_i  P_i · mean(R_i)

── Volatilidad esperada del portafolio ──────────────────────
  Cov(R)  = matriz de covarianza de R  → [M x M]

  Var(Rp) = P · Cov(R) · Pᵀ          → escalar
  σ(Rp)   = √(P · Cov(R) · Pᵀ)

Nota: al trabajar con retornos simulados (escenarios),
  mean() y cov() se calculan sobre las N filas de R,
  capturando implícitamente la correlación entre activos.
""")

# ──────────────────────────────────────────────
# 2.2  CÓMPUTO NUMÉRICO
# ──────────────────────────────────────────────
print("="*60)
print("2.2  CÓMPUTO NUMÉRICO")
print("="*60)

# Vectores de medias y matriz de covarianza
mu    = R.mean(axis=0)          # [55]
cov_R = np.cov(R, rowvar=False) # [55 x 55]

def portfolio_stats(P, label):
    """Calcula retorno esperado y volatilidad de un portafolio."""
    ret    = P @ mu                          # E[Rp]
    var    = P @ cov_R @ P                   # Var(Rp)
    vol    = np.sqrt(var)                    # σ(Rp)
    sharpe = ret / vol if vol > 0 else np.nan
    return ret, vol, sharpe

ret_p1, vol_p1, sr_p1 = portfolio_stats(P1, "P1")
ret_p2, vol_p2, sr_p2 = portfolio_stats(P2, "P2")

print(f"""
┌─────────────────────────────────────────────────────┐
│         Resultados de Portafolios P1 y P2           │
├──────────────────────┬──────────────┬───────────────┤
│ Métrica              │      P1      │      P2       │
├──────────────────────┼──────────────┼───────────────┤
│ Retorno esperado     │  {ret_p1:>10.4%}  │  {ret_p2:>11.4%}  │
│ Volatilidad esperada │  {vol_p1:>10.4%}  │  {vol_p2:>11.4%}  │
│ Ratio Sharpe (rf=0)  │  {sr_p1:>10.4f}  │  {sr_p2:>11.4f}  │
└──────────────────────┴──────────────┴───────────────┘
""")

# ── Detalle de pesos no nulos ──
print("── Pesos no nulos P1 ─────────────────────────────────")
for i, w in enumerate(P1):
    if w != 0:
        print(f"  Activo {i+1:2d}:  w={w:.3f}   μ={mu[i]:.4%}   contribución={w*mu[i]:.4%}")

print(f"\n── Pesos no nulos P2 ─────────────────────────────────")
for i, w in enumerate(P2):
    if w != 0:
        print(f"  Activo {i+1:2d}:  w={w:.3f}   μ={mu[i]:.4%}   contribución={w*mu[i]:.4%}")

# ── Contribución marginal al riesgo ──
def marginal_risk(P, cov, label):
    total_vol = np.sqrt(P @ cov @ P)
    mcr = (cov @ P) / total_vol          # vector de riesgo marginal
    risk_contrib = P * mcr               # contribución al riesgo de cada activo
    return risk_contrib

print(f"\n── Contribución al riesgo (Risk Budgeting) ───────────")
rc1 = marginal_risk(P1, cov_R, "P1")
rc2 = marginal_risk(P2, cov_R, "P2")

print(f"\n{'Activo':<10} {'P1 contrib. riesgo':>20} {'P2 contrib. riesgo':>20}")
print("-"*52)
for i in range(55):
    if abs(rc1[i]) > 1e-6 or abs(rc2[i]) > 1e-6:
        print(f"Activo {i+1:<4} {rc1[i]:>20.4%} {rc2[i]:>20.4%}")

print(f"\n{'TOTAL':<10} {rc1.sum():>20.4%} {rc2.sum():>20.4%}")

print("\n✓ Análisis completo.")

"""
2.3 Haga la siguiente pregunta a Claude y a Gemini: “¿Cómo afecta una subida de tasas
de la Fed al retorno y volatilidad esperados de un portafolio de renta fija y uno de renta
variable?”. Compare los outputs en una tabla según profundidad, precisión, tono y
estructura, y concluya cual modelo escogería y por qué.


| Criterio    | Claude                                                                                                  | Gemini                                                                                   |
| ----------- | ------------------------------------------------------------------------------------------------------- | ---------------------------------------------------------------------------------------- |
| Profundidad | Muy alta: Detalles cuantitativos (duración 10% por 1%), spreads, betas, escenarios (stagflation).       | Alta: Explica canales (WACC, P/E), fórmula descuento; accesible pero menos cuantitativo. |
| Precisión   | Excelente: Alineada con teoría (duración, VIX histórico, correlación 60/40 en 2022). | Excelente: Correcta en mecanismos indirectos y fórmula; nota largo plazo precisa.pnc+1   |
| Tono        | Técnico, profesional y objetivo.                                                                        | Amigable, metafórico ("gravedad"), engaging con pregunta final.                          |
| Estructura  | Superior: Headers, listas, tabla comparativa, resumida clara.                                           | Buena: Numerada, tabla resumen, fórmula LaTeX, pero más narrativa.                       |

Me parecio Claude superior ya que me dio un analisis mas profundo, claro y con una mejor estructura, ademas de que su tono fue mas profesional y tecnico, mientras que Gemini aunque fue muy bueno, se sintio un poco mas informal y menos detallado en los aspectos cuantitativos, lo cual es clave para un analista de inversiones.

"""
