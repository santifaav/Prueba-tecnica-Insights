"""
Ejercicio 3 — Caso de Automatización
Prueba Técnica Analista Jr de Inversiones — Insights WM
=======================================================
Motor de decisiones para solicitudes de retiro.
Produce:
  - decisions_db.xlsx   : todas las solicitudes con decisión, reason_code y severidad
  - review_queue.xlsx   : solo HOLDs ordenados por severidad desc
"""

import pandas as pd
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constantes ──────────────────────────────────────────────────────────────
BUFFER_USD       = 50
DUP_WINDOW_MIN   = 15

SEVERITY = {
    "KYC_NOT_VERIFIED":                    100,
    "ACCOUNT_NOT_ACTIVE":                  95,
    "UNWHITELISTED_HIGH_AML":              90,
    "INVALID_AMOUNT":                      85,
    "DUPLICATE_REQUEST":                   70,
    "INSUFFICIENT_SETTLED_AFTER_BUFFER":   65,
    "INSUFFICIENT_AVAILABLE_AFTER_BUFFER": 55,
    "DEST_CHANGED_RECENTLY":               45,
    "URGENT_RISK_TIER":                    35,
}

# ── Carga de datos ───────────────────────────────────────────────────────────
FILE = "withdrawals.xlsx"

req = pd.read_excel(FILE, sheet_name="withdrawal_requests")
snap = pd.read_excel(FILE, sheet_name="account_snapshot")
dest = pd.read_excel(FILE, sheet_name="destination_registry")

# Parsear fechas con timezone-awareness
req["created_at"]       = pd.to_datetime(req["created_at"], utc=True)
snap["as_of"]           = pd.to_datetime(snap["as_of"], utc=True)
dest["last_changed_at"] = pd.to_datetime(dest["last_changed_at"], utc=True)

# Merge request con account snapshot y destination registry
df = req.merge(snap, on="account_id", how="left", suffixes=("", "_snap"))
df = df.merge(dest[["destination_id","is_whitelisted","last_changed_at","client_id"]],
              on="destination_id", how="left", suffixes=("", "_dest"))

# Renombrar columna de client_id del dest para no confundir
df.rename(columns={"client_id_dest": "dest_client_id"}, inplace=True)

# ── Función de decisión ──────────────────────────────────────────────────────
def evaluate(row, seen_requests):
    """
    Retorna (decision, reason_code, severity)
    Prioridad: REJECT → HOLD → APPROVE
    """
    reasons_reject = []
    reasons_hold   = []

    # ── REJECT checks ────────────────────────────────────────────────────────

    # 1. account_status ≠ active
    if pd.isna(row["account_status"]) or row["account_status"] != "active":
        reasons_reject.append("ACCOUNT_NOT_ACTIVE")

    # 2. kyc_status ≠ verified
    if pd.isna(row["kyc_status"]) or row["kyc_status"] != "verified":
        reasons_reject.append("KYC_NOT_VERIFIED")

    # 3. amount ≤ 0
    if pd.isna(row["amount"]) or row["amount"] <= 0:
        reasons_reject.append("INVALID_AMOUNT")

    # 4. Duplicado: mismo account_id + amount + destination_id dentro de 15 min
    key = (row["account_id"], row["amount"], row["destination_id"])
    ts  = row["created_at"]
    if key in seen_requests:
        for prev_ts in seen_requests[key]:
            diff_min = abs((ts - prev_ts).total_seconds()) / 60
            if diff_min <= DUP_WINDOW_MIN:
                reasons_reject.append("DUPLICATE_REQUEST")
                break
    # Registrar esta solicitud (solo si no es ya un duplicado marcado)
    seen_requests.setdefault(key, []).append(ts)

    # 5. aml_risk_tier = high y is_whitelisted = false
    if row.get("aml_risk_tier") == "high" and row.get("is_whitelisted") == False:
        reasons_reject.append("UNWHITELISTED_HIGH_AML")

    # Si hay razones de rechazo, retornar la de mayor severidad
    if reasons_reject:
        top = max(reasons_reject, key=lambda r: SEVERITY[r])
        return "REJECT", top, SEVERITY[top]

    # ── HOLD checks ──────────────────────────────────────────────────────────

    # 1. available_cash - amount < BUFFER_USD
    if not pd.isna(row["available_cash"]) and (row["available_cash"] - row["amount"] < BUFFER_USD):
        reasons_hold.append("INSUFFICIENT_AVAILABLE_AFTER_BUFFER")

    # 2. settled_cash - amount < BUFFER_USD
    if not pd.isna(row["settled_cash"]) and (row["settled_cash"] - row["amount"] < BUFFER_USD):
        reasons_hold.append("INSUFFICIENT_SETTLED_AFTER_BUFFER")

    # 3. last_changed_at del destino dentro de RECENT_DEST_DAYS vs as_of snapshot
    if not pd.isna(row["last_changed_at"]) and not pd.isna(row["as_of"]):
        days_since = (row["as_of"] - row["last_changed_at"]).total_seconds() / 86400
        if days_since < 8:
            reasons_hold.append("DEST_CHANGED_RECENTLY")

    # 4. requested_speed = urgent y aml_risk_tier ∈ {medium, high}
    if row.get("requested_speed") == "urgent" and row.get("aml_risk_tier") in {"medium", "high"}:
        reasons_hold.append("URGENT_RISK_TIER")

    if reasons_hold:
        top = max(reasons_hold, key=lambda r: SEVERITY[r])
        return "HOLD", top, SEVERITY[top]

    return "APPROVE", "OK", 0


# ── Aplicar lógica ───────────────────────────────────────────────────────────
# Ordenar por created_at para correcta detección de duplicados en ventana de tiempo
df = df.sort_values("created_at").reset_index(drop=True)

seen_requests = {}
results = []
for _, row in df.iterrows():
    decision, reason, sev = evaluate(row, seen_requests)
    results.append({"request_id": row["request_id"],
                    "decision":   decision,
                    "reason_code": reason,
                    "severity":   sev})

res_df = pd.DataFrame(results)

# ── Merge para enriquecer la base de decisiones ──────────────────────────────
decisions_df = req.merge(res_df, on="request_id")
decisions_df["created_at"] = decisions_df["created_at"].dt.strftime("%Y-%m-%d %H:%M:%S")

# Reordenar columnas relevantes
cols_order = ["request_id","created_at","client_id","account_id",
              "amount","currency","destination_id","requested_speed",
              "decision","reason_code","severity"]
decisions_df = decisions_df[cols_order]

# Separar HOLDs para review queue
review_df = decisions_df[decisions_df["decision"] == "HOLD"].sort_values("severity", ascending=False)

# ── Resumen en consola ────────────────────────────────────────────────────────
counts = decisions_df["decision"].value_counts()
print("=" * 55)
print("   RESUMEN DE DECISIONES")
print("=" * 55)
print(f"  Total solicitudes  : {len(decisions_df)}")
print(f"  ✅  APPROVE        : {counts.get('APPROVE', 0)}")
print(f"  🔶  HOLD           : {counts.get('HOLD', 0)}")
print(f"  ❌  REJECT         : {counts.get('REJECT', 0)}")
print()
print("  Distribución de reason_codes:")
rc_counts = decisions_df[decisions_df["decision"] != "APPROVE"]["reason_code"].value_counts()
for rc, cnt in rc_counts.items():
    print(f"    {rc:<42} {cnt:>3}")
print("=" * 55)

# ── Helpers de estilo ─────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F3864")   # azul oscuro
APPROVE_FILL = PatternFill("solid", start_color="C6EFCE")   # verde
HOLD_FILL    = PatternFill("solid", start_color="FFEB9C")   # amarillo
REJECT_FILL  = PatternFill("solid", start_color="FFC7CE")   # rojo
WHITE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

def style_sheet(ws, df, decision_col_idx=None, decision_fills=None):
    """Aplica estilo profesional a una hoja."""
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.font  = WHITE_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        decision = None
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font   = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c_idx != 2 else LEFT
            if df.columns[c_idx-1] == "decision":
                decision = val
        # Color por fila según decisión
        if decision and decision_fills:
            fill = decision_fills.get(decision)
            if fill:
                for c_idx in range(1, len(df.columns)+1):
                    ws.cell(row=r_idx, column=c_idx).fill = fill

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── Generar decisions_db.xlsx ─────────────────────────────────────────────────
wb1 = Workbook()

# ── Hoja 1: Todas las decisiones ─────────────────────────────────────────────
ws1 = wb1.active
ws1.title = "Decisions DB"
style_sheet(ws1, decisions_df,
            decision_fills={"APPROVE": APPROVE_FILL,
                            "HOLD":    HOLD_FILL,
                            "REJECT":  REJECT_FILL})

# ── Hoja 2: Solo REJECTs ──────────────────────────────────────────────────────
reject_df = decisions_df[decisions_df["decision"] == "REJECT"].sort_values("severity", ascending=False)
ws2 = wb1.create_sheet("Rejects")
style_sheet(ws2, reject_df,
            decision_fills={"REJECT": REJECT_FILL})

# ── Hoja 3: Solo APPROVEs ─────────────────────────────────────────────────────
approve_df = decisions_df[decisions_df["decision"] == "APPROVE"]
ws3 = wb1.create_sheet("Approvals")
style_sheet(ws3, approve_df,
            decision_fills={"APPROVE": APPROVE_FILL})

# ── Hoja 4: Dashboard resumen ──────────────────────────────────────────────────
ws4 = wb1.create_sheet("Dashboard")
ws4.column_dimensions["A"].width = 35
ws4.column_dimensions["B"].width = 15
ws4.column_dimensions["C"].width = 15

# Título
ws4["A1"] = "DECISIONS DASHBOARD — Insights WM"
ws4["A1"].font = Font(name="Arial", bold=True, size=14, color="1F3864")
ws4.merge_cells("A1:C1")
ws4["A1"].alignment = CENTER

ws4["A3"] = "DECISIÓN";    ws4["B3"] = "CANTIDAD";  ws4["C3"] = "% TOTAL"
for c in ["A3","B3","C3"]:
    ws4[c].font = WHITE_FONT; ws4[c].fill = HEADER_FILL; ws4[c].alignment = CENTER

total = len(decisions_df)
for i, (dec, fill) in enumerate([("APPROVE",APPROVE_FILL),("HOLD",HOLD_FILL),("REJECT",REJECT_FILL)], 4):
    cnt = counts.get(dec, 0)
    ws4[f"A{i}"] = dec;           ws4[f"A{i}"].fill = fill
    ws4[f"B{i}"] = cnt;           ws4[f"B{i}"].fill = fill
    ws4[f"C{i}"] = f"{cnt/total:.1%}"; ws4[f"C{i}"].fill = fill
    for c in [f"A{i}",f"B{i}",f"C{i}"]:
        ws4[c].font = BOLD_FONT; ws4[c].alignment = CENTER; ws4[c].border = THIN_BORDER

ws4["A8"] = "REASON CODE"; ws4["B8"] = "DECISIÓN"; ws4["C8"] = "CANTIDAD"
for c in ["A8","B8","C8"]:
    ws4[c].font = WHITE_FONT; ws4[c].fill = HEADER_FILL; ws4[c].alignment = CENTER; ws4[c].border = THIN_BORDER

rc_full = decisions_df[decisions_df["decision"] != "APPROVE"][["decision","reason_code"]]\
            .groupby(["reason_code","decision"]).size().reset_index(name="count")\
            .sort_values("count", ascending=False)
for i2, row2 in enumerate(rc_full.itertuples(index=False), 9):
    fill = HOLD_FILL if row2.decision == "HOLD" else REJECT_FILL
    ws4[f"A{i2}"] = row2.reason_code; ws4[f"A{i2}"].fill = fill; ws4[f"A{i2}"].border = THIN_BORDER
    ws4[f"B{i2}"] = row2.decision;    ws4[f"B{i2}"].fill = fill; ws4[f"B{i2}"].border = THIN_BORDER
    ws4[f"C{i2}"] = row2.count;       ws4[f"C{i2}"].fill = fill; ws4[f"C{i2}"].border = THIN_BORDER
    for c in [f"A{i2}",f"B{i2}",f"C{i2}"]:
        ws4[c].font = BODY_FONT; ws4[c].alignment = CENTER

wb1.save("decisions_db.xlsx")
print("✅  decisions_db.xlsx generado")

# ── Generar review_queue.xlsx ─────────────────────────────────────────────────
wb2 = Workbook()
ws_r = wb2.active
ws_r.title = "Review Queue"

style_sheet(ws_r, review_df,
            decision_fills={"HOLD": HOLD_FILL})

wb2.save("review_queue.xlsx")
print("✅  review_queue.xlsx generado")
print(f"\n   HOLDs en cola de revisión: {len(review_df)}")
print(f"   Severidad máxima: {review_df['severity'].max()}")
print(f"   Severidad mínima: {review_df['severity'].min()}")
