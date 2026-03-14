"""
Agente de Retiros — Insights WM
================================
Ejercicio 3 (Bono): Agentización escalable con Claude API

Flujo:
  1. Motor de reglas determinista  →  APPROVE / REJECT / HOLD
  2. HOLDs  →  Claude API enriquece con narrativa + prioridad
  3. Outputs:  decisions_db.xlsx  |  review_queue.xlsx  |  audit_log.json

Requisitos:
  pip install anthropic pandas openpyxl

Uso:
  python withdrawal_agent.py
  python withdrawal_agent.py --input withdrawals.xlsx --output ./outputs
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constantes ────────────────────────────────────────────────────────────────
BUFFER_USD       = 50
DUP_WINDOW_MIN   = 15
MODEL            = "claude-sonnet-4-6"

SEVERITY = {
    "KYC_NOT_VERIFIED":                    100,
    "ACCOUNT_NOT_ACTIVE":                  95,
    "UNWHITELISTED_HIGH_AML":              90,
    "INVALID_AMOUNT":                      85,
    "DUPLICATE_REQUEST":                   70,
    "INSUFFICIENT_SETTLED_AFTER_BUFFER":   65,
    "INSUFFICIENT_AVAILABLE_AFTER_BUFFER": 55,
    "DEST_CHANGED_RECENTLY":               45,
    "URGENT_RISK_TIER":                    35,
}

REASON_DESCRIPTIONS = {
    "KYC_NOT_VERIFIED":                    "KYC del cliente no está verificado",
    "ACCOUNT_NOT_ACTIVE":                  "La cuenta no está activa (frozen/closed)",
    "UNWHITELISTED_HIGH_AML":              "Destino no whitelisted con AML alto",
    "INVALID_AMOUNT":                      "Monto inválido (≤ 0)",
    "DUPLICATE_REQUEST":                   "Solicitud duplicada en ventana de 15 min",
    "INSUFFICIENT_SETTLED_AFTER_BUFFER":   "Fondos settled insuficientes tras buffer",
    "INSUFFICIENT_AVAILABLE_AFTER_BUFFER": "Fondos disponibles insuficientes tras buffer",
    "DEST_CHANGED_RECENTLY":               "Destino modificado en los últimos 7 días",
    "URGENT_RISK_TIER":                    "Retiro urgente con AML medium/high",
}

# ── Estilos Excel ─────────────────────────────────────────────────────────────
H_FILL   = PatternFill("solid", start_color="1F3864")
APR_FILL = PatternFill("solid", start_color="C6EFCE")
HLD_FILL = PatternFill("solid", start_color="FFEB9C")
REJ_FILL = PatternFill("solid", start_color="FFC7CE")
W_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
B_FONT   = Font(name="Arial", size=10)
BD_FONT  = Font(name="Arial", bold=True, size=10)
CENTER   = Alignment(horizontal="center", vertical="center")
BORDER   = Border(*[Side(style="thin")]*0,
                  left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))

FILLS = {"APPROVE": APR_FILL, "HOLD": HLD_FILL, "REJECT": REJ_FILL}


# ══════════════════════════════════════════════════════════════════════════════
# 1. MOTOR DE REGLAS (determinista — sin LLM)
# ══════════════════════════════════════════════════════════════════════════════
def evaluate_request(row: pd.Series, seen: dict) -> tuple[str, str, int]:
    """
    Retorna (decision, reason_code, severity).
    Prioridad: REJECT > HOLD > APPROVE.
    """
    rej, hld = [], []

    # ── REJECT ────────────────────────────────────────────────────────────────
    if pd.isna(row.get("account_status")) or row["account_status"] != "active":
        rej.append("ACCOUNT_NOT_ACTIVE")
    if pd.isna(row.get("kyc_status")) or row["kyc_status"] != "verified":
        rej.append("KYC_NOT_VERIFIED")
    if pd.isna(row.get("amount")) or row["amount"] <= 0:
        rej.append("INVALID_AMOUNT")

    key = (row["account_id"], row["amount"], row["destination_id"])
    ts  = row["created_at"]
    if key in seen:
        for prev in seen[key]:
            if abs((ts - prev).total_seconds()) / 60 <= DUP_WINDOW_MIN:
                rej.append("DUPLICATE_REQUEST")
                break
    seen.setdefault(key, []).append(ts)

    if row.get("aml_risk_tier") == "high" and row.get("is_whitelisted") is False:
        rej.append("UNWHITELISTED_HIGH_AML")

    if rej:
        top = max(rej, key=lambda r: SEVERITY[r])
        return "REJECT", top, SEVERITY[top]

    # ── HOLD ──────────────────────────────────────────────────────────────────
    av  = row.get("available_cash", float("nan"))
    stl = row.get("settled_cash",   float("nan"))
    amt = row["amount"]

    if not pd.isna(av)  and (av  - amt < BUFFER_USD): hld.append("INSUFFICIENT_AVAILABLE_AFTER_BUFFER")
    if not pd.isna(stl) and (stl - amt < BUFFER_USD): hld.append("INSUFFICIENT_SETTLED_AFTER_BUFFER")

    if not pd.isna(row.get("last_changed_at")) and not pd.isna(row.get("as_of")):
        days = (row["as_of"] - row["last_changed_at"]).total_seconds() / 86400
        if days < 8:
            hld.append("DEST_CHANGED_RECENTLY")

    if row.get("requested_speed") == "urgent" and row.get("aml_risk_tier") in {"medium", "high"}:
        hld.append("URGENT_RISK_TIER")

    if hld:
        top = max(hld, key=lambda r: SEVERITY[r])
        return "HOLD", top, SEVERITY[top]

    return "APPROVE", "OK", 0


# ══════════════════════════════════════════════════════════════════════════════
# 2. AGENTE LLM — enriquecimiento de HOLDs con Claude API
# ══════════════════════════════════════════════════════════════════════════════
SYSTEM_PROMPT = """Eres un analista de operaciones senior en Insights WM, una firma de
wealth management. Tu rol es revisar solicitudes de retiro que han sido marcadas en HOLD
por el sistema de reglas automatizado y producir:

1. Un resumen ejecutivo claro (2-3 oraciones) explicando por qué está en revisión.
2. Una recomendación de acción: APPROVE_AFTER_REVIEW o ESCALATE.
3. Una justificación de riesgo (1-2 oraciones).
4. Un score de urgencia adicional del 1 al 5 (1 = baja, 5 = crítica).

Responde SIEMPRE en el siguiente formato JSON (sin markdown, sin texto extra):
{
  "summary": "...",
  "recommendation": "APPROVE_AFTER_REVIEW" | "ESCALATE",
  "risk_justification": "...",
  "urgency_score": 1-5
}

Sé conciso, preciso y orientado al riesgo. Usa el contexto del cliente (AML tier,
velocidad, monto, historial de destino) para priorizar correctamente."""


def enrich_hold_with_llm(client, row: pd.Series,
                          reason_code: str, severity: int) -> dict:
    """
    Llama a Claude API para enriquecer un HOLD con narrativa y recomendación.
    Retorna dict con campos: summary, recommendation, risk_justification, urgency_score.
    """
    context = f"""
Solicitud de retiro en HOLD — Insights WM

request_id      : {row.get('request_id', 'N/A')}
account_id      : {row.get('account_id', 'N/A')}
client_id       : {row.get('client_id', 'N/A')}
amount          : USD {row.get('amount', 0):,.2f}
destination_id  : {row.get('destination_id', 'N/A')}
requested_speed : {row.get('requested_speed', 'N/A')}
channel         : {row.get('channel', 'N/A')}
created_at      : {row.get('created_at', 'N/A')}

Estado de la cuenta:
  account_status : {row.get('account_status', 'N/A')}
  kyc_status     : {row.get('kyc_status', 'N/A')}
  aml_risk_tier  : {row.get('aml_risk_tier', 'N/A')}
  available_cash : USD {row.get('available_cash', 0):,.2f}
  settled_cash   : USD {row.get('settled_cash', 0):,.2f}

Destino:
  is_whitelisted  : {row.get('is_whitelisted', 'N/A')}
  last_changed_at : {row.get('last_changed_at', 'N/A')}
  snapshot as_of  : {row.get('as_of', 'N/A')}

Razón del HOLD (motor de reglas):
  reason_code : {reason_code}
  descripción : {REASON_DESCRIPTIONS.get(reason_code, 'N/A')}
  severidad   : {severity}
""".strip()

    try:
        message = client.messages.create(
            model=MODEL,
            max_tokens=400,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": context}]
        )
        text = message.content[0].text.strip()
        # Limpiar posibles bloques markdown
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        return json.loads(text)
    except (json.JSONDecodeError, anthropic.APIError, Exception) as e:
        return {
            "summary":            f"Error al procesar con LLM: {str(e)[:80]}",
            "recommendation":     "ESCALATE",
            "risk_justification": "Revisión manual requerida — análisis LLM no disponible.",
            "urgency_score":      3
        }


# ══════════════════════════════════════════════════════════════════════════════
# 3. GENERADOR DE EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def style_ws(ws, df: pd.DataFrame):
    """Aplica estilo profesional con colores por decisión."""
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, col.upper())
        c.font, c.fill, c.alignment, c.border = W_FONT, H_FILL, CENTER, BORDER

    for ri, row in enumerate(df.itertuples(index=False), 2):
        dec = None
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.font, c.border = B_FONT, BORDER
            c.alignment = Alignment(horizontal="left" if ci == 2 else "center",
                                    vertical="center", wrap_text=(ci == len(df.columns)))
            if df.columns[ci-1] == "decision":
                dec = val
        if dec and dec in FILLS:
            for ci in range(1, len(df.columns)+1):
                ws.cell(ri, ci).fill = FILLS[dec]

    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 55)
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def build_outputs(decisions_df: pd.DataFrame, review_df: pd.DataFrame,
                  out_dir: Path):
    """Genera decisions_db.xlsx y review_queue.xlsx."""
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── decisions_db.xlsx ──────────────────────────────────────────────────
    wb = Workbook()
    ws_all = wb.active; ws_all.title = "All Decisions"
    style_ws(ws_all, decisions_df)

    for name, mask in [("Approvals", "APPROVE"), ("Holds", "HOLD"), ("Rejects", "REJECT")]:
        ws = wb.create_sheet(name)
        sub = decisions_df[decisions_df["decision"] == mask]
        if not sub.empty:
            style_ws(ws, sub)

    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.column_dimensions["A"].width = 38
    ws_dash.column_dimensions["B"].width = 12
    ws_dash.column_dimensions["C"].width = 12
    ws_dash["A1"] = "Decisions Dashboard — Insights WM"
    ws_dash["A1"].font = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws_dash.merge_cells("A1:C1")
    ws_dash["A1"].alignment = CENTER
    total = len(decisions_df)
    counts = decisions_df["decision"].value_counts()
    for i, h in enumerate(["DECISIÓN","CANTIDAD","% TOTAL"], 1):
        c = ws_dash.cell(3, i, h)
        c.font, c.fill, c.alignment, c.border = W_FONT, H_FILL, CENTER, BORDER
    for ri, (dec, fill) in enumerate([("APPROVE",APR_FILL),("HOLD",HLD_FILL),("REJECT",REJ_FILL)], 4):
        cnt = counts.get(dec, 0)
        for ci, val in enumerate([dec, cnt, f"{cnt/total:.1%}"], 1):
            c = ws_dash.cell(ri, ci, val)
            c.font, c.fill, c.alignment, c.border = BD_FONT, fill, CENTER, BORDER

    wb.save(out_dir / "decisions_db.xlsx")
    print(f"  ✅  decisions_db.xlsx  → {out_dir}/decisions_db.xlsx")

    # ── review_queue.xlsx ──────────────────────────────────────────────────
    wb2 = Workbook()
    ws_r = wb2.active; ws_r.title = "Review Queue"
    if not review_df.empty:
        style_ws(ws_r, review_df)
    wb2.save(out_dir / "review_queue.xlsx")
    print(f"  ✅  review_queue.xlsx  → {out_dir}/review_queue.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# 4. PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
def run(input_file: str, output_dir: str):
    print("\n" + "═"*58)
    print("  Agente de Retiros — Insights WM")
    print("═"*58)

    # ── Carga ──────────────────────────────────────────────────────────────
    print("\n[1/4] Cargando datos...")
    req  = pd.read_excel(input_file, sheet_name="withdrawal_requests")
    snap = pd.read_excel(input_file, sheet_name="account_snapshot")
    dest = pd.read_excel(input_file, sheet_name="destination_registry")

    req["created_at"]       = pd.to_datetime(req["created_at"], utc=True)
    snap["as_of"]           = pd.to_datetime(snap["as_of"], utc=True)
    dest["last_changed_at"] = pd.to_datetime(dest["last_changed_at"], utc=True)

    df = (req
          .merge(snap, on="account_id", how="left", suffixes=("","_snap"))
          .merge(dest[["destination_id","is_whitelisted","last_changed_at","client_id"]],
                 on="destination_id", how="left", suffixes=("","_dest"))
          .sort_values("created_at").reset_index(drop=True))

    print(f"  {len(df)} solicitudes cargadas.")

    # ── Motor de reglas ────────────────────────────────────────────────────
    print("\n[2/4] Aplicando motor de reglas...")
    seen, raw = {}, []
    for _, row in df.iterrows():
        dec, reason, sev = evaluate_request(row, seen)
        raw.append({"request_id": row["request_id"],
                    "decision":    dec,
                    "reason_code": reason,
                    "severity":    sev})
    res = pd.DataFrame(raw)

    counts = res["decision"].value_counts()
    print(f"  APPROVE: {counts.get('APPROVE',0)}  |  "
          f"HOLD: {counts.get('HOLD',0)}  |  "
          f"REJECT: {counts.get('REJECT',0)}")

    # ── Enriquecimiento LLM de HOLDs ───────────────────────────────────────
    print("\n[3/4] Enriqueciendo HOLDs con Claude API...")
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    llm_enrichments = []

    if not api_key or not ANTHROPIC_AVAILABLE:
        msg = "ANTHROPIC_API_KEY no configurada" if not api_key else "librería anthropic no instalada (pip install anthropic)"
        print(f"  ⚠️  {msg} — HOLDs sin narrativa LLM.")
    else:
        client = anthropic.Anthropic(api_key=api_key)
        hold_ids = res[res["decision"] == "HOLD"]["request_id"].tolist()
        hold_rows = df[df["request_id"].isin(hold_ids)].copy()
        hold_rows = hold_rows.merge(res[["request_id","reason_code","severity"]],
                                    on="request_id")

        total_holds = len(hold_rows)
        for i, (_, hrow) in enumerate(hold_rows.iterrows(), 1):
            print(f"  → [{i}/{total_holds}] {hrow['request_id']}  "
                  f"({hrow['reason_code']})  ", end="", flush=True)
            enrichment = enrich_hold_with_llm(
                client, hrow, hrow["reason_code"], hrow["severity"]
            )
            enrichment["request_id"] = hrow["request_id"]
            llm_enrichments.append(enrichment)
            print(f"✓  urgency={enrichment['urgency_score']}  "
                  f"rec={enrichment['recommendation']}")
            time.sleep(0.3)  # rate-limit gentil

    # ── Construir DataFrames finales ───────────────────────────────────────
    cols_base = ["request_id","created_at","client_id","account_id",
                 "amount","currency","destination_id","requested_speed",
                 "decision","reason_code","severity"]
    decisions_df = req.merge(res, on="request_id")[cols_base].copy()
    decisions_df["created_at"] = decisions_df["created_at"].dt.strftime("%Y-%m-%d %H:%M:%S")

    # Review queue = HOLDs + narrativa LLM
    review_cols = cols_base + ["llm_summary","llm_recommendation",
                                "llm_risk_justification","llm_urgency_score"]
    review_df = decisions_df[decisions_df["decision"] == "HOLD"].copy()

    if llm_enrichments:
        enr_df = pd.DataFrame(llm_enrichments).rename(columns={
            "summary":           "llm_summary",
            "recommendation":    "llm_recommendation",
            "risk_justification":"llm_risk_justification",
            "urgency_score":     "llm_urgency_score",
        })
        review_df = review_df.merge(enr_df[["request_id","llm_summary",
                                             "llm_recommendation",
                                             "llm_risk_justification",
                                             "llm_urgency_score"]],
                                     on="request_id", how="left")
    else:
        for col in ["llm_summary","llm_recommendation",
                    "llm_risk_justification","llm_urgency_score"]:
            review_df[col] = "N/A"

    review_df = review_df.sort_values("severity", ascending=False).reset_index(drop=True)

    # ── Audit log (JSON) ───────────────────────────────────────────────────
    audit = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "input_file":   input_file,
        "model":        MODEL,
        "totals": {k: int(v) for k, v in counts.items()},
        "decisions": decisions_df.to_dict(orient="records"),
        "llm_enrichments": llm_enrichments,
    }

    # ── Outputs ────────────────────────────────────────────────────────────
    print("\n[4/4] Generando archivos de salida...")
    out_dir = Path(output_dir)
    build_outputs(decisions_df, review_df, out_dir)

    audit_path = out_dir / "audit_log.json"
    with open(audit_path, "w", encoding="utf-8") as f:
        json.dump(audit, f, ensure_ascii=False, indent=2, default=str)
    print(f"  ✅  audit_log.json     → {audit_path}")

    print("\n" + "═"*58)
    print("  Pipeline completado exitosamente.")
    print("═"*58 + "\n")
    return decisions_df, review_df


# ══════════════════════════════════════════════════════════════════════════════
# 5. BONUS: Modo interactivo — simula revisión de un HOLD por consola
# ══════════════════════════════════════════════════════════════════════════════
def interactive_review(review_df: pd.DataFrame):
    """
    Permite al analista revisar HOLDs uno por uno desde la consola,
    simulando el panel de revisión humana.
    """
    if review_df.empty:
        print("No hay solicitudes en HOLD para revisar.")
        return

    print("\n" + "═"*58)
    print("  MODO REVISIÓN INTERACTIVA — HOLDs pendientes")
    print("═"*58)

    for i, row in review_df.iterrows():
        print(f"\n[{i+1}/{len(review_df)}] {row['request_id']}")
        print(f"  Monto      : USD {row['amount']:,.2f}")
        print(f"  Razón      : {row['reason_code']} (severity={row['severity']})")
        print(f"  Velocidad  : {row['requested_speed']}")
        if row.get("llm_summary") not in (None, "N/A"):
            print(f"\n  🤖 Análisis Claude:")
            print(f"     {row['llm_summary']}")
            print(f"     Recomendación : {row['llm_recommendation']}")
            print(f"     Urgencia      : {row['llm_urgency_score']}/5")
            print(f"     Riesgo        : {row['llm_risk_justification']}")

        action = input("\n  ¿Acción? [a=approve / r=reject / s=skip / q=quit]: ").strip().lower()
        if action == "q":
            print("  Revisión interrumpida.")
            break
        elif action == "a":
            print(f"  ✅  {row['request_id']} → APROBADO manualmente.")
        elif action == "r":
            print(f"  ❌  {row['request_id']} → RECHAZADO manualmente.")
        else:
            print(f"  ⏭   {row['request_id']} → omitido.")


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Agente de retiros — Insights WM")
    parser.add_argument("--input",       default="withdrawals.xlsx",
                        help="Ruta al archivo withdrawals.xlsx")
    parser.add_argument("--output",      default="./outputs",
                        help="Directorio de salida")
    parser.add_argument("--interactive", action="store_true",
                        help="Activa modo de revisión interactiva por consola")
    args = parser.parse_args()

    decisions_df, review_df = run(args.input, args.output)

    if args.interactive:
        interactive_review(review_df)
